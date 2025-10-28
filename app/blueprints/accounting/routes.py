from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, abort, current_app, jsonify
from flask_babel import gettext as _
from flask_login import login_required
from ...security import role_required
from ...extensions import db, mail
from ...models import (
    Vehicle,
    Shipment,
    Invoice,
    InvoiceItem,
    Payment,
    InternationalCost,
    BillOfLading,
    Customer,
    Setting,
    VehicleShipment,
    Account,
    ExchangeRate,
    JournalEntry,
    JournalLine,
    OperationalExpense,
    CustomerDeposit,
    ClientAccountStructure,
    VehicleAccountStructure,
    Auction,
)
from ...utils_pdf import render_invoice_pdf, render_bol_pdf, render_vehicle_statement_pdf
import os
from flask_mail import Message
from datetime import datetime
from decimal import Decimal

acct_bp = Blueprint("acct", __name__, template_folder="templates/accounting")

def _normalize_number_string(value: object) -> str:
    """Normalize user-entered numeric strings including Arabic-Indic digits and separators.

    Converts Arabic digits to ASCII, removes thousands separators, and ensures a '.' decimal.
    Returns a safe string that float() can parse.
    """
    if value is None:
        return "0"
    try:
        s = str(value)
    except Exception:
        return "0"
    s = s.strip()
    if not s:
        return "0"
    # Map Arabic-Indic and Eastern Arabic-Indic digits; normalize separators
    translate_map = str.maketrans({
        "٠": "0", "١": "1", "٢": "2", "٣": "3", "٤": "4",
        "٥": "5", "٦": "6", "٧": "7", "٨": "8", "٩": "9",
        "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
        "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
        # Decimal and thousands separators (Arabic)
        "٫": ".",  # U+066B ARABIC DECIMAL SEPARATOR
        "٬": "",   # U+066C ARABIC THOUSANDS SEPARATOR
        "،": "",   # U+060C ARABIC COMMA (treat as thousands separator)
        # Non‑breaking space
        "\u00A0": "",
    })
    s = s.translate(translate_map)
    # Handle standard commas: treat single comma (and no dot present) as decimal; otherwise remove
    if "," in s:
        if "." not in s and s.count(",") == 1:
            left, right = s.split(",", 1)
            if 1 <= len(right) <= 3:
                s = f"{left}.{right}"
            else:
                s = s.replace(",", "")
        else:
            s = s.replace(",", "")
    # Remove regular spaces
    s = s.replace(" ", "")
    # Keep only digits, one leading '-', and dots
    cleaned = []
    for ch in s:
        if ch.isdigit() or ch in ".-":
            cleaned.append(ch)
    s = "".join(cleaned)
    if s in {"", "-", ".", "-."}:
        return "0"
    return s

def _parse_number_input(value: object) -> float:
    try:
        return float(_normalize_number_string(value))
    except Exception:
        return 0.0

def _get_account(code: str) -> Account | None:
    try:
        return db.session.query(Account).filter(Account.code == code).first()
    except Exception:
        return None


def _ensure_client_accounts(customer: Customer) -> ClientAccountStructure:
    """Create per-client sub-accounts if missing and return mapping row.

    Accounts created under these base parents by convention:
      - Bank remains global A100
      - Client Deposits parent L200 (we create L200-C{ID})
      - Auction Payments clearing under A150 (optional A150-C{ID})
      - Service Revenue under R300 (R300-C{ID})
      - Logistics Expense under E200 (E200-C{ID})
      - Accounts Receivable under A300 (A300-C{ID})
    """
    cas = db.session.query(ClientAccountStructure).filter_by(customer_id=customer.id).first()
    if cas:
        return cas
    cid = int(customer.id)
    code_suffix = f"C{cid:05d}"
    # Prepare derived codes
    dep_code = f"L200-{code_suffix}"
    auc_code = f"A150-{code_suffix}"
    srv_code = f"R300-{code_suffix}"
    log_code = f"E200-{code_suffix}"
    ar_code = f"A300-{code_suffix}"

    def ensure_account(code: str, name: str, typ: str) -> Account:
        acc = db.session.query(Account).filter(Account.code == code).first()
        if not acc:
            acc = Account(code=code, name=name, type=typ, client_id=customer.id)
            db.session.add(acc)
            db.session.flush()
        return acc

    display_name = (customer.company_name or customer.full_name or f"Client {customer.id}").strip()
    ensure_account(dep_code, f"{display_name} Deposit", "LIABILITY")
    ensure_account(auc_code, f"{display_name} Auction Clearing", "ASSET")
    ensure_account(srv_code, f"{display_name} Service Revenue", "REVENUE")
    ensure_account(log_code, f"{display_name} Logistics Expense", "EXPENSE")
    ensure_account(ar_code, f"{display_name} Receivable", "ASSET")

    cas = ClientAccountStructure(
        customer_id=customer.id,
        deposit_account_code=dep_code,
        auction_account_code=auc_code,
        service_revenue_account_code=srv_code,
        logistics_expense_account_code=log_code,
        receivable_account_code=ar_code,
        currency_code="OMR",
    )
    db.session.add(cas)
    db.session.flush()
    return cas


def _ensure_vehicle_accounts(vehicle: Vehicle) -> VehicleAccountStructure:
    """Create per-vehicle sub-accounts if missing and return mapping row.

    Accounts are created under conventional parent groupings and tagged with vehicle_id:
      - Deposit (L200-V{vehicle_id})
      - Auction Clearing (A150-V{vehicle_id})
      - Freight Expense (E200-V{vehicle_id})
      - Customs Expense (E220-V{vehicle_id})
      - Commission Revenue (R300-V{vehicle_id})
      - Storage Expense (E230-V{vehicle_id})
    """
    vas = db.session.query(VehicleAccountStructure).filter_by(vehicle_id=vehicle.id).first()
    if vas:
        return vas
    vid = int(vehicle.id)
    vin = (vehicle.vin or f"V{vid:06d}").strip()
    owner_id = getattr(vehicle, 'owner_customer_id', None)
    # Derived codes
    dep_code = f"L200-V{vid:06d}"
    auc_code = f"A150-V{vid:06d}"
    frt_code = f"E200-V{vid:06d}"
    cst_code = f"E220-V{vid:06d}"
    com_code = f"R300-V{vid:06d}"
    str_code = f"E230-V{vid:06d}"

    def ensure_account(code: str, name: str, typ: str):
        acc = db.session.query(Account).filter(Account.code == code).first()
        if not acc:
            acc = Account(code=code, name=name, type=typ, client_id=owner_id, vehicle_id=vid)
            db.session.add(acc)
            db.session.flush()
        return acc

    label = vin
    ensure_account(dep_code, f"{label} Deposit", "LIABILITY")
    ensure_account(auc_code, f"{label} Auction", "ASSET")
    ensure_account(frt_code, f"{label} Freight", "EXPENSE")
    ensure_account(cst_code, f"{label} Customs", "EXPENSE")
    ensure_account(com_code, f"{label} Commission", "REVENUE")
    ensure_account(str_code, f"{label} Storage", "EXPENSE")

    vas = VehicleAccountStructure(
        vehicle_id=vid,
        client_id=owner_id,
        deposit_account_code=dep_code,
        auction_account_code=auc_code,
        freight_account_code=frt_code,
        customs_account_code=cst_code,
        commission_account_code=com_code,
        storage_account_code=str_code,
        currency_code='OMR',
    )
    db.session.add(vas)
    db.session.flush()
    return vas


def _get_vehicle_account_code(vehicle_id: int | None, kind: str, default_code: str) -> str:
    """Return the per-vehicle account code for kind if mapping exists; fallback to default_code.

    kind: 'deposit' | 'auction' | 'freight' | 'customs' | 'commission' | 'storage'
    """
    if not vehicle_id:
        return default_code
    vas = db.session.query(VehicleAccountStructure).filter_by(vehicle_id=vehicle_id).first()
    if not vas:
        v = db.session.get(Vehicle, vehicle_id)
        if not v:
            return default_code
        vas = _ensure_vehicle_accounts(v)
    return {
        'deposit': vas.deposit_account_code,
        'auction': vas.auction_account_code,
        'freight': vas.freight_account_code,
        'customs': vas.customs_account_code,
        'commission': vas.commission_account_code,
        'storage': vas.storage_account_code,
    }.get(kind, default_code)


def create_vehicle_chart(vehicle_id: int, client_id: int | None = None) -> VehicleAccountStructure | None:
    """Public API to ensure a vehicle's sub-ledger accounts exist.

    Returns the created or existing VehicleAccountStructure.
    """
    if not vehicle_id:
        return None
    v = db.session.get(Vehicle, vehicle_id)
    if not v:
        return None
    # Optionally set or update client linkage on the structure
    vas = _ensure_vehicle_accounts(v)
    if client_id and getattr(vas, 'client_id', None) != client_id:
        try:
            vas.client_id = client_id
            db.session.flush()
        except Exception:
            pass
    return vas

def _get_client_account_code(customer_id: int | None, kind: str, default_code: str) -> str:
    """Return the per-client account code for kind if customer has mapping; fallback to default_code.

    kind: 'deposit' | 'auction' | 'service' | 'logistics' | 'receivable'
    """
    if not customer_id:
        return default_code
    cas = db.session.query(ClientAccountStructure).filter_by(customer_id=customer_id).first()
    if not cas:
        cust = db.session.get(Customer, customer_id)
        if not cust:
            return default_code
        cas = _ensure_client_accounts(cust)
    return {
        'deposit': cas.deposit_account_code,
        'auction': cas.auction_account_code or default_code,
        'service': cas.service_revenue_account_code,
        'logistics': cas.logistics_expense_account_code,
        'receivable': cas.receivable_account_code,
    }.get(kind, default_code)


def create_client_chart(client_id: int) -> ClientAccountStructure | None:
    """Public API to ensure a client's sub-ledger accounts exist.

    Returns the created or existing ClientAccountStructure.
    """
    if not client_id:
        return None
    cust = db.session.get(Customer, client_id)
    if not cust:
        return None
    return _ensure_client_accounts(cust)

def _post_journal(description: str, reference: str | None, lines: list[tuple[str, float, float]],
                  customer_id: int | None = None, vehicle_id: int | None = None,
                  auction_id: int | None = None, invoice_id: int | None = None,
                  is_client_fund: bool = False, status: str = 'approved', notes: str | None = None):
    """Create a balanced journal entry from (account_code, debit, credit) lines.
    Amounts are in OMR.
    """
    from ...models import Setting
    # Enforce lock period
    try:
        settings_row = db.session.query(Setting).first()
        if settings_row and getattr(settings_row, 'books_locked_until', None):
            from datetime import datetime
            if datetime.utcnow() <= settings_row.books_locked_until:
                status = 'pending'  # queue for approval if within locked period
    except Exception:
        pass

    entry = JournalEntry(
        description=description,
        reference=reference,
        customer_id=customer_id,
        vehicle_id=vehicle_id,
        auction_id=auction_id,
        invoice_id=invoice_id,
        is_client_fund=bool(is_client_fund),
        status=status or 'approved',
        notes=notes,
    )
    db.session.add(entry)
    db.session.flush()
    total_debit = 0
    total_credit = 0
    for code, dr, cr in lines:
        acc = _get_account(code)
        if not acc:
            # Failsafe: skip line if account missing
            continue
        dr_amt = float(dr or 0)
        cr_amt = float(cr or 0)
        total_debit += dr_amt
        total_credit += cr_amt
        db.session.add(JournalLine(entry_id=entry.id, account_id=acc.id, debit=dr_amt, credit=cr_amt, currency_code='OMR'))
    # Do not enforce balance hard to avoid blocking UI; rely on tests/admin checks
    return entry


# ---- Manual Journals (basic CRUD & listing) ----
@acct_bp.route('/journals')
@role_required('accountant', 'admin')
def journals_list():
    q = db.session.query(JournalEntry).order_by(JournalEntry.entry_date.desc(), JournalEntry.id.desc())
    entries = q.limit(200).all()
    return render_template('accounting/journals_list.html', entries=entries)


@acct_bp.route('/journals/new', methods=['GET','POST'])
@role_required('accountant', 'admin')
def journals_new():
    if request.method == 'POST':
        description = (request.form.get('description') or '').strip()
        reference = (request.form.get('reference') or '').strip() or None
        is_client_fund = (request.form.get('is_client_fund') == 'on')
        # Expect parallel lists of codes/debits/credits
        codes = request.form.getlist('code')
        drs = request.form.getlist('debit')
        crs = request.form.getlist('credit')
        lines: list[tuple[str, float, float]] = []
        for c, d, r in zip(codes, drs, crs):
            c = (c or '').strip()
            if not c:
                continue
            lines.append((c, _parse_number_input(d), _parse_number_input(r)))
        _post_journal(description=description or 'Manual journal', reference=reference, lines=lines, is_client_fund=is_client_fund)
        try:
            db.session.commit(); flash(_('Journal posted'), 'success')
        except Exception:
            db.session.rollback(); flash(_('Failed to post journal'), 'danger')
        return redirect(url_for('acct.journals_list'))
    return render_template('accounting/journals_form.html')


@acct_bp.route('/journals/<int:entry_id>/approve', methods=['POST'])
@role_required('accountant', 'admin')
def journals_approve(entry_id: int):
    je = db.session.get(JournalEntry, entry_id)
    if not je:
        flash(_('Not found'), 'danger'); return redirect(url_for('acct.journals_list'))
    je.status = 'approved'
    je.approved_at = datetime.utcnow()
    try:
        from flask_login import current_user
        je.approved_by_user_id = getattr(current_user, 'id', None)
    except Exception:
        pass
    try:
        db.session.commit(); flash(_('Journal approved'), 'success')
    except Exception:
        db.session.rollback(); flash(_('Failed to approve journal'), 'danger')
    return redirect(url_for('acct.journals_list'))


@acct_bp.route('/journals/<int:entry_id>/delete', methods=['POST'])
@role_required('accountant', 'admin')
def journals_delete(entry_id: int):
    je = db.session.get(JournalEntry, entry_id)
    if not je:
        flash(_('Not found'), 'danger'); return redirect(url_for('acct.journals_list'))
    db.session.delete(je)
    try:
        db.session.commit(); flash(_('Journal deleted'), 'success')
    except Exception:
        db.session.rollback(); flash(_('Failed to delete journal'), 'danger')
    return redirect(url_for('acct.journals_list'))

# ---- Stage 1: Customer Deposit (Security) ----
def record_customer_deposit(customer_id: int, amount_omr: float, method: str | None = None,
                            reference: str | None = None, vehicle_id: int | None = None,
                            auction_id: int | None = None) -> CustomerDeposit:
    dep = CustomerDeposit(customer_id=customer_id, vehicle_id=vehicle_id, auction_id=auction_id,
                          amount_omr=amount_omr, method=method, reference=reference, status='held')
    db.session.add(dep)
    # Journal: Dr Bank (A100) / Cr Customer Deposits (client sub-account under L200)
    dep_code = _get_vehicle_account_code(vehicle_id, 'deposit', _get_client_account_code(customer_id, 'deposit', 'L200'))
    _post_journal(
        description='Customer deposit received', reference=reference,
        lines=[('A100', amount_omr, 0.0), (dep_code, 0.0, amount_omr)],
        customer_id=customer_id, vehicle_id=vehicle_id, auction_id=auction_id,
        is_client_fund=True,
    )
    return dep

def refund_customer_deposit(deposit_id: int):
    dep = db.session.get(CustomerDeposit, deposit_id)
    if not dep or dep.status != 'held':
        return False
    dep.status = 'refunded'
    dep.refunded_at = datetime.utcnow()
    # Journal reversal: Dr Customer Deposits / Cr Bank
    amt = float(dep.amount_omr or 0)
    dep_code = _get_vehicle_account_code(dep.vehicle_id, 'deposit', _get_client_account_code(dep.customer_id, 'deposit', 'L200'))
    _post_journal(
        description='Customer deposit refunded', reference=dep.reference,
        lines=[(dep_code, amt, 0.0), ('A100', 0.0, amt)],
        customer_id=dep.customer_id, vehicle_id=dep.vehicle_id, auction_id=dep.auction_id,
        is_client_fund=True,
    )
    return True


# ---- Client fund applications & commissions ----
def pay_auction_from_client_fund(customer_id: int, amount_omr: float, reference: str | None = None,
                                 vehicle_id: int | None = None, auction_id: int | None = None):
    """Use held client deposit to pay auction: Dr Client Deposits / Cr Bank.

    Flagged as client fund so excluded from P&L.
    """
    if float(amount_omr or 0) <= 0:
        return None
    dep_code = _get_vehicle_account_code(vehicle_id, 'deposit', _get_client_account_code(customer_id, 'deposit', 'L200'))
    return _post_journal(
        description='Auction payment from client funds', reference=reference,
        lines=[(dep_code, float(amount_omr), 0.0), ('A100', 0.0, float(amount_omr))],
        customer_id=customer_id, vehicle_id=vehicle_id, auction_id=auction_id,
        is_client_fund=True,
    )


def record_commission_from_deposit(customer_id: int, amount_omr: float, reference: str | None = None,
                                   vehicle_id: int | None = None, invoice_id: int | None = None):
    """Recognize commission by deducting from client deposit: Dr Client Deposits / Cr Revenue (R300)."""
    if float(amount_omr or 0) <= 0:
        return None
    dep_code = _get_vehicle_account_code(vehicle_id, 'deposit', _get_client_account_code(customer_id, 'deposit', 'L200'))
    rev_code = _get_vehicle_account_code(vehicle_id, 'commission', _get_client_account_code(customer_id, 'service', 'R300'))
    return _post_journal(
        description='Commission deducted from client deposit', reference=reference,
        lines=[(dep_code, float(amount_omr), 0.0), (rev_code, 0.0, float(amount_omr))],
        customer_id=customer_id, vehicle_id=vehicle_id, invoice_id=invoice_id,
        is_client_fund=True,
    )

# ---- Stage 2: Car Invoice after winning auction ----
def create_car_invoice(customer_id: int, vehicle_id: int, price_omr: float,
                       optional_fees_omr: float = 0.0, deposit_applied_omr: float = 0.0) -> int:
    inv = Invoice(invoice_number=f"CAR-{int(datetime.utcnow().timestamp())}", customer_id=customer_id,
                  vehicle_id=vehicle_id, invoice_type='CAR', status='Draft', total_omr=0)
    db.session.add(inv)
    db.session.flush()
    items_total = Decimal('0')
    db.session.add(InvoiceItem(invoice_id=inv.id, vehicle_id=vehicle_id, description='Car price', amount_omr=price_omr))
    items_total += Decimal(str(price_omr))
    if optional_fees_omr and float(optional_fees_omr) > 0:
        db.session.add(InvoiceItem(invoice_id=inv.id, vehicle_id=vehicle_id, description='Optional fees', amount_omr=optional_fees_omr))
        items_total += Decimal(str(optional_fees_omr))
    inv.total_omr = items_total
    # Defer revenue recognition until actual payment is recorded.
    # Keep invoice as Unpaid so it doesn't count towards profits.
    if float(items_total) > 0:
        inv.status = 'Unpaid'
    return inv.id


# ---- Stage 3: Purchase at auction and shipping costs ----
def record_vehicle_purchase(vehicle_id: int, auction_id: int | None, purchase_price_usd: float,
                            paid_from_bank: bool = True):
    veh = db.session.get(Vehicle, vehicle_id)
    omr_rate = Decimal(str(current_app.config.get('OMR_EXCHANGE_RATE', 0.385)))
    amount_omr = Decimal(str(purchase_price_usd or 0)) * omr_rate
    # Inventory capitalization (as asset) at OMR
    _post_journal(
        description='Vehicle purchased at auction', reference=getattr(veh, 'vin', None),
        lines=[
            ('A200', float(amount_omr), 0.0),
            ('A100', 0.0, float(amount_omr)) if paid_from_bank else ('L210', 0.0, float(amount_omr)),
        ],
        vehicle_id=vehicle_id, auction_id=auction_id,
        is_client_fund=not paid_from_bank,
    )
    return float(amount_omr)

def record_operational_cost(vehicle_id: int | None, auction_id: int | None, category: str,
                            amount_value: float, currency: str = 'OMR', description: str | None = None,
                            supplier: str | None = None, paid_from_bank: bool = True) -> int:
    # Convert to OMR if needed
    rate_val = Decimal('1')
    rate_row = None
    if currency and currency.upper() != 'OMR':
        rate_row = db.session.query(ExchangeRate).order_by(ExchangeRate.effective_at.desc()).first()
        rate_val = Decimal(str(rate_row.rate if rate_row else current_app.config.get('OMR_EXCHANGE_RATE', 0.385)))
    amount_omr = Decimal(str(amount_value or 0)) * (rate_val if currency.upper() != 'OMR' else Decimal('1'))

    exp = OperationalExpense(
        vehicle_id=vehicle_id, auction_id=auction_id, category=category,
        original_amount=amount_value, original_currency=(currency or 'OMR').upper(), amount_omr=float(amount_omr),
        exchange_rate_id=(rate_row.id if rate_row else None), description=description, supplier=supplier,
        paid=bool(paid_from_bank), paid_at=datetime.utcnow() if paid_from_bank else None,
    )
    db.session.add(exp)

    # Journal: Dr Operational Expenses / Cr Bank (if paid)
    if float(amount_omr) > 0 and paid_from_bank:
        # Try to attribute the expense to a client via vehicle owner or auction customer
        customer_id = None
        try:
            if vehicle_id:
                v = db.session.get(Vehicle, int(vehicle_id))
                customer_id = getattr(v, 'owner_customer_id', None)
        except Exception:
            customer_id = customer_id
        if not customer_id and auction_id:
            try:
                a = db.session.get(Auction, int(auction_id))
                customer_id = getattr(a, 'customer_id', None)
            except Exception:
                customer_id = customer_id
        # Map category to per-vehicle account when possible
        kind = 'freight'
        cat_norm = (category or '').lower()
        if 'custom' in cat_norm:
            kind = 'customs'
        elif 'storage' in cat_norm or 'warehouse' in cat_norm:
            kind = 'storage'
        else:
            kind = 'freight'
        exp_code_default = 'E200'
        exp_code = _get_vehicle_account_code(vehicle_id, kind, _get_client_account_code(customer_id, 'logistics', exp_code_default))
        _post_journal(
            description=f'Operational expense - {category}', reference=description,
            lines=[(exp_code, float(amount_omr), 0.0), ('A100', 0.0, float(amount_omr))],
            customer_id=customer_id, vehicle_id=vehicle_id, auction_id=auction_id,
        )

    return exp.id if getattr(exp, 'id', None) else 0


# ---- Stage 4: Shipping invoice to customer with fines ----
def create_shipping_invoice(customer_id: int, vehicle_id: int, shipping_cost_omr: float,
                            fines_usd: float = 0.0) -> int:
    # Determine rate for fines conversion
    rate_row = db.session.query(ExchangeRate).order_by(ExchangeRate.effective_at.desc()).first()
    rate_val = Decimal(str(rate_row.rate if rate_row else current_app.config.get('OMR_EXCHANGE_RATE', 0.385)))
    fines_omr = Decimal(str(fines_usd or 0)) * rate_val

    inv = Invoice(
        invoice_number=f"SHP-{int(datetime.utcnow().timestamp())}",
        customer_id=customer_id,
        vehicle_id=vehicle_id,
        invoice_type='SHIPPING',
        status='Draft',
        exchange_rate_id=(rate_row.id if rate_row else None),
        total_omr=0,
    )
    db.session.add(inv)
    db.session.flush()
    total = Decimal('0')
    if shipping_cost_omr and float(shipping_cost_omr) > 0:
        db.session.add(InvoiceItem(invoice_id=inv.id, vehicle_id=vehicle_id, description='Shipping cost', amount_omr=shipping_cost_omr))
        total += Decimal(str(shipping_cost_omr))
    if fines_omr and float(fines_omr) > 0:
        db.session.add(InvoiceItem(invoice_id=inv.id, vehicle_id=vehicle_id, description='Fines (converted to OMR)', amount_omr=float(fines_omr)))
        total += fines_omr
    inv.total_omr = total

    # Assume immediate collection for simplicity: 
    # Dr Bank (total), Cr Fines Revenue (fines_omr), Cr Operational Expenses (shipping_cost_omr) to offset prior expense
    if float(total) > 0:
        lines = [('A100', float(total), 0.0)]
        if float(fines_omr) > 0:
            lines.append((_get_vehicle_account_code(vehicle_id, 'commission', _get_client_account_code(customer_id, 'service', 'R300')), 0.0, float(fines_omr)))
        if float(shipping_cost_omr or 0) > 0:
            lines.append((_get_vehicle_account_code(vehicle_id, 'freight', _get_client_account_code(customer_id, 'logistics', 'E200')), 0.0, float(shipping_cost_omr)))
        _post_journal(
            description='Shipping invoice payment', reference=inv.invoice_number,
            lines=lines,
            customer_id=customer_id, vehicle_id=vehicle_id, invoice_id=inv.id,
            is_client_fund=False,
        )
        inv.status = 'Paid'
    return inv.id
@acct_bp.route("/dashboard")
@role_required("accountant", "admin")
def dashboard():
    # summary metrics
    counts = {
        "invoices": db.session.query(Invoice).count(),
        "vehicles_priced": db.session.query(InternationalCost).count(),
    }

    usd_to_omr = float(current_app.config.get('OMR_EXCHANGE_RATE', 0.385))
    # Ensure numeric operations are done with consistent types to avoid Decimal*float TypeError
    freight_usd_sum = db.session.query(db.func.coalesce(db.func.sum(Shipment.cost_freight_usd), 0)).scalar() or 0
    auction_fees_usd_sum = db.session.query(db.func.coalesce(db.func.sum(InternationalCost.auction_fees_usd), 0)).scalar() or 0
    expenses_omr = (float(freight_usd_sum) + float(auction_fees_usd_sum)) * usd_to_omr
    # Treat CAR invoices as pass-through costs (expenses), not revenue
    car_paid_total = float(
        db.session.query(db.func.coalesce(db.func.sum(Invoice.total_omr), 0))
        .filter(Invoice.status == 'Paid', Invoice.invoice_type == 'CAR')
        .scalar()
        or 0
    )
    # Revenue should include service fees only (R* accounts excluding client funds)
    rev_total = db.session.query(db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0)).\
        join(Account, JournalLine.account_id == Account.id).\
        join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
        filter(Account.code.like('R%'), JournalEntry.is_client_fund.is_(False)).scalar() or 0
    totals = {
        "revenue_omr": float(rev_total),
        "expenses_omr": expenses_omr + car_paid_total,
    }
    totals["net_omr"] = totals["revenue_omr"] - totals["expenses_omr"]

    # monthly revenue series (last 12 months)
    now = datetime.utcnow()
    months = []
    revenue_series = []
    for i in range(12):
        month_start = datetime(now.year - (1 if now.month - i <= 0 else 0), ((now.month - i - 1) % 12) + 1, 1)
        if month_start.month == 12:
            month_end = datetime(month_start.year + 1, 1, 1)
        else:
            month_end = datetime(month_start.year, month_start.month + 1, 1)
        # Sum monthly revenue from GL (R* accounts), excluding client funds
        total = (
            db.session.query(db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0))
            .join(Account, JournalLine.account_id == Account.id)
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(
                JournalEntry.entry_date >= month_start,
                JournalEntry.entry_date < month_end,
                Account.code.like('R%'),
                JournalEntry.is_client_fund.is_(False),
            )
            .scalar()
            or 0
        )
        months.append(month_start.strftime('%b'))
        revenue_series.append(float(total))
    months = list(reversed(months))
    revenue_series = list(reversed(revenue_series))

    # expenses by category (freight, customs, vat, local transport, misc)
    exp = {
        "Freight": float(db.session.query(db.func.coalesce(db.func.sum(Shipment.cost_freight_usd), 0)).scalar() or 0) * usd_to_omr,
        "Customs": float(db.session.query(db.func.coalesce(db.func.sum(InternationalCost.customs_omr), 0)).scalar() or 0),
        "VAT": float(db.session.query(db.func.coalesce(db.func.sum(InternationalCost.vat_omr), 0)).scalar() or 0),
        "Local Transport": float(db.session.query(db.func.coalesce(db.func.sum(InternationalCost.local_transport_omr), 0)).scalar() or 0),
        "Misc": float(db.session.query(db.func.coalesce(db.func.sum(InternationalCost.misc_omr), 0)).scalar() or 0),
        "Car Price": float(
            db.session.query(db.func.coalesce(db.func.sum(Invoice.total_omr), 0))
            .filter(Invoice.status == 'Paid', Invoice.invoice_type == 'CAR')
            .scalar()
            or 0
        ),
    }

    # KPI: outstanding client deposits (L200* credit balance)
    client_deposits = db.session.query(db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0)).\
        join(Account, JournalLine.account_id == Account.id).\
        filter(Account.code.like('L200%')).scalar() or 0
    totals["client_deposits_omr"] = float(client_deposits)

    return render_template("accounting/dashboard.html", counts=counts, totals=totals, chart={
        "months": months, "revenue": revenue_series, "exp_labels": list(exp.keys()), "exp_values": list(exp.values())
    })


# International Costs Management
@acct_bp.route('/costs')
@role_required('accountant', 'admin')
def costs_list():
    q = db.session.query(Vehicle).order_by(Vehicle.created_at.desc())
    vehicles = q.limit(50).all()
    return render_template('accounting/costs_list.html', vehicles=vehicles)


@acct_bp.route('/costs/<int:vehicle_id>', methods=['GET','POST'])
@role_required('accountant', 'admin')
def costs_edit(vehicle_id: int):
    vehicle = db.session.get(Vehicle, vehicle_id)
    if not vehicle:
        flash(_('Vehicle not found'), 'danger')
        return redirect(url_for('acct.costs_list'))
    cost = db.session.query(InternationalCost).filter_by(vehicle_id=vehicle.id).first()
    if request.method == 'POST':
        def f(name: str) -> float:
            return _parse_number_input(request.form.get(name))
        if not cost:
            cost = InternationalCost(vehicle_id=vehicle.id)
            db.session.add(cost)
        cost.freight_usd = f('freight_usd')
        cost.insurance_usd = f('insurance_usd')
        cost.auction_fees_usd = f('auction_fees_usd')
        cost.customs_omr = f('customs_omr')
        cost.vat_omr = f('vat_omr')
        cost.local_transport_omr = f('local_transport_omr')
        cost.misc_omr = f('misc_omr')
        try:
            db.session.commit()
            flash(_('Costs saved'), 'success')
            return redirect(url_for('acct.costs_list'))
        except Exception:
            db.session.rollback()
            flash(_('Failed to save costs'), 'danger')
    return render_template('accounting/costs_edit.html', vehicle=vehicle, cost=cost)


# Invoices CRUD
@acct_bp.route('/invoices')
@role_required('accountant', 'admin')
def invoices_list():
    invoices = db.session.query(Invoice).order_by(Invoice.created_at.desc()).all()
    return render_template('accounting/invoices_list.html', invoices=invoices)


@acct_bp.route('/invoices/new', methods=['GET','POST'])
@role_required('accountant', 'admin')
def invoices_new():
    customers = db.session.query(Customer).order_by(Customer.company_name.asc()).all()
    vehicles = db.session.query(Vehicle).order_by(Vehicle.created_at.desc()).limit(100).all()
    if request.method == 'POST':
        customer_id = request.form.get('customer_id')
        items = []
        descriptions = request.form.getlist('item_description')
        amounts = request.form.getlist('item_amount')
        for d, a in zip(descriptions, amounts):
            if d.strip():
                try:
                    items.append((d.strip(), float(a or 0)))
                except Exception:
                    items.append((d.strip(), 0.0))
        inv = Invoice(invoice_number=f"INV-{int(datetime.utcnow().timestamp())}", customer_id=int(customer_id) if customer_id else None, status='Draft', total_omr=0)
        db.session.add(inv)
        db.session.flush()
        total = Decimal('0')
        for d, a in items:
            db.session.add(InvoiceItem(invoice_id=inv.id, description=d, amount_omr=a))
            total += Decimal(str(a))
        inv.total_omr = total
        try:
            db.session.commit()
            flash(_('Invoice created'), 'success')
            return redirect(url_for('acct.invoices_edit', invoice_id=inv.id))
        except Exception:
            db.session.rollback()
            flash(_('Failed to create invoice'), 'danger')
    return render_template('accounting/invoices_form.html', customers=customers, vehicles=vehicles)


@acct_bp.route('/invoices/<int:invoice_id>/edit', methods=['GET','POST'])
@role_required('accountant', 'admin')
def invoices_edit(invoice_id: int):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        flash(_('Invoice not found'), 'danger')
        return redirect(url_for('acct.invoices_list'))
    customers = db.session.query(Customer).order_by(Customer.company_name.asc()).all()
    if request.method == 'POST':
        status = request.form.get('status') or 'Draft'
        invoice.status = status
        # replace items
        db.session.query(InvoiceItem).filter_by(invoice_id=invoice.id).delete()
        descriptions = request.form.getlist('item_description')
        amounts = request.form.getlist('item_amount')
        total = Decimal('0')
        for d, a in zip(descriptions, amounts):
            if d.strip():
                val = Decimal(str(a or 0))
                db.session.add(InvoiceItem(invoice_id=invoice.id, description=d.strip(), amount_omr=val))
                total += val
        invoice.total_omr = total
        # If invoice is for services (not CAR) and marked Unpaid, recognize AR and Revenue once
        if invoice.invoice_type != 'CAR' and str(status).strip().lower() == 'unpaid':
            # Has revenue already been recognized for this invoice?
            exists = (
                db.session.query(JournalEntry)
                .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
                .join(Account, JournalLine.account_id == Account.id)
                .filter(JournalEntry.invoice_id == invoice.id, Account.code.like('R%'))
                .first()
            )
            if not exists and float(invoice.total_omr or 0) > 0:
                # Dr AR (client) / Cr Revenue (client)
                ar_code = _get_client_account_code(invoice.customer_id, 'receivable', 'A300')
                rev_code = _get_client_account_code(invoice.customer_id, 'service', 'R300')
                _post_journal(
                    description='Service invoice issued', reference=invoice.invoice_number,
                    lines=[(ar_code, float(invoice.total_omr), 0.0), (rev_code, 0.0, float(invoice.total_omr))],
                    customer_id=invoice.customer_id, invoice_id=invoice.id, is_client_fund=False,
                )
        try:
            db.session.commit()
            flash(_('Invoice updated'), 'success')
        except Exception:
            db.session.rollback()
            flash(_('Failed to update invoice'), 'danger')
    return render_template('accounting/invoices_edit.html', invoice=invoice, customers=customers)


@acct_bp.route('/invoices/<int:invoice_id>/delete', methods=['POST'])
@role_required('accountant', 'admin')
def invoices_delete(invoice_id: int):
    inv = db.session.get(Invoice, invoice_id)
    if not inv:
        flash(_('Not found'), 'danger')
        return redirect(url_for('acct.invoices_list'))
    db.session.delete(inv)
    try:
        db.session.commit()
        flash(_('Invoice deleted'), 'success')
    except Exception:
        db.session.rollback()
        flash(_('Failed to delete invoice'), 'danger')
    return redirect(url_for('acct.invoices_list'))


@acct_bp.route('/invoices/<int:invoice_id>/export')
@role_required('accountant', 'admin')
def invoices_export(invoice_id: int):
    inv = db.session.get(Invoice, invoice_id)
    if not inv:
        abort(404)
    items = db.session.query(InvoiceItem).filter_by(invoice_id=inv.id).all()
    path = render_invoice_pdf(inv, items)
    inv.pdf_path = path
    db.session.commit()
    return send_file(path, as_attachment=True, download_name=f"{inv.invoice_number}.pdf")


@acct_bp.route('/invoices/<int:invoice_id>/export.xlsx')
@role_required('accountant', 'admin')
def invoices_export_xlsx(invoice_id: int):
    from openpyxl import Workbook
    inv = db.session.get(Invoice, invoice_id)
    if not inv:
        abort(404)
    items = db.session.query(InvoiceItem).filter_by(invoice_id=inv.id).all()
    wb = Workbook(); ws = wb.active; ws.title = inv.invoice_number or 'Invoice'
    ws.append(['Invoice #', inv.invoice_number])
    ws.append(['Date', inv.created_at.strftime('%Y-%m-%d') if inv.created_at else ''])
    ws.append(['Client', inv.customer.display_name if inv.customer else '-'])
    ws.append([])
    ws.append(['Description', 'Amount (OMR)'])
    for it in items:
        ws.append([it.description, float(it.amount_omr or 0)])
    ws.append([])
    ws.append(['Total', float(inv.total_omr or 0)])
    from io import BytesIO
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"{inv.invoice_number}.xlsx")

@acct_bp.route('/invoices/<int:invoice_id>/email', methods=['POST'])
@role_required('accountant', 'admin')
def invoices_email(invoice_id: int):
    inv = db.session.get(Invoice, invoice_id)
    if not inv or not inv.customer or not inv.customer.user:
        flash(_('Missing customer email'), 'danger')
        return redirect(url_for('acct.invoices_list'))
    email = inv.customer.user.email
    # ensure pdf exists
    items = db.session.query(InvoiceItem).filter_by(invoice_id=inv.id).all()
    path = inv.pdf_path
    if not path or not os.path.isfile(path):
        try:
            current_app.logger.warning("Invoice PDF missing at %s; regenerating", path)
        except Exception:
            pass
        path = render_invoice_pdf(inv, items)
        inv.pdf_path = path
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
    try:
        msg = Message(subject=_('Invoice %(n)s', n=inv.invoice_number), recipients=[email])
        msg.body = _('Please find attached invoice %(n)s.', n=inv.invoice_number)
        with open(path, 'rb') as f:
            msg.attach(filename=f"{inv.invoice_number}.pdf", content_type='application/pdf', data=f.read())
        mail.send(msg)
        flash(_('Email sent'), 'success')
    except Exception:
        flash(_('Failed to send email'), 'danger')
    return redirect(url_for('acct.invoices_edit', invoice_id=inv.id))


# Payments
@acct_bp.route('/payments')
@role_required('accountant', 'admin')
def payments_list():
    payments = db.session.query(Payment).order_by(Payment.created_at.desc()).limit(100).all()
    invoices = db.session.query(Invoice).order_by(Invoice.created_at.desc()).all()
    return render_template('accounting/payments_list.html', payments=payments, invoices=invoices)


@acct_bp.get('/api/invoices/by_vin')
@role_required('accountant', 'admin')
def api_invoice_by_vin():
    """Return the most relevant invoice for a given VIN (chassis).

    Preference order:
      1) The invoice with the largest outstanding balance (> 0)
      2) Otherwise, the latest invoice for that vehicle
    """
    vin = (request.args.get('vin') or '').strip()
    if not vin:
        return jsonify({'error': 'vin_required'}), 400
    try:
        veh = (
            db.session.query(Vehicle)
            .filter(db.func.upper(Vehicle.vin) == db.func.upper(vin))
            .first()
        )
    except Exception:
        veh = None
    if not veh:
        return jsonify({'error': 'vehicle_not_found'}), 404

    invs = (
        db.session.query(Invoice)
        .filter(Invoice.vehicle_id == veh.id)
        .order_by(Invoice.created_at.desc(), Invoice.id.desc())
        .all()
    )
    if not invs:
        return jsonify({'error': 'no_invoices_for_vehicle'}), 404

    chosen = None
    chosen_balance = -1.0
    for inv in invs:
        try:
            total = float(inv.total_omr or 0)
        except Exception:
            total = 0.0
        try:
            paid = float(inv.paid_total() or 0)
        except Exception:
            paid = 0.0
        balance = max(0.0, total - paid)
        if balance > chosen_balance:
            chosen = inv
            chosen_balance = balance
    inv = chosen or invs[0]

    try:
        total = float(inv.total_omr or 0)
    except Exception:
        total = 0.0
    try:
        paid = float(inv.paid_total() or 0)
    except Exception:
        paid = 0.0
    balance = max(0.0, total - paid)

    items = [
        {
            'description': (it.description or '-'),
            'amount_omr': float(it.amount_omr or 0),
        }
        for it in (inv.items or [])
    ]

    out = {
        'invoice_id': inv.id,
        'invoice_number': inv.invoice_number,
        'invoice_type': inv.invoice_type,
        'status': inv.status,
        'total_omr': total,
        'paid_omr': paid,
        'balance_omr': balance,
        'vehicle': {
            'id': veh.id,
            'vin': veh.vin,
            'status': veh.status,
        },
        'customer': {
            'id': inv.customer_id,
            'name': (inv.customer.display_name if inv.customer else '-'),
        },
        'items': items,
    }
    return jsonify(out)


@acct_bp.route('/payments/new', methods=['POST'])
@role_required('accountant', 'admin')
def payments_new():
    invoice_id = request.form.get('invoice_id')
    amount = request.form.get('amount')
    method = request.form.get('method')
    reference = request.form.get('reference')
    vin_input = (request.form.get('vin') or '').strip().upper()

    # Enforce VIN-only flow: VIN must be provided
    if not vin_input:
        flash(_('VIN is required'), 'danger')
        return redirect(url_for('acct.payments_list'))

    # If invoice_id is missing, try resolving from VIN
    inv = db.session.get(Invoice, int(invoice_id)) if invoice_id else None
    if not inv:
        try:
            veh = (
                db.session.query(Vehicle)
                .filter(db.func.upper(Vehicle.vin) == vin_input)
                .first()
            )
        except Exception:
            veh = None
        if veh:
            cand = (
                db.session.query(Invoice)
                .filter(Invoice.vehicle_id == veh.id)
                .order_by(Invoice.created_at.desc(), Invoice.id.desc())
                .first()
            )
            inv = cand
    if not inv:
        flash(_('Invalid invoice'), 'danger')
        return redirect(url_for('acct.payments_list'))
    try:
        amt = Decimal(str(amount or 0))
    except Exception:
        amt = Decimal('0')
    # Try to link vehicle and customer using VIN; otherwise fall back to invoice linkage
    vehicle_id = None
    customer_id = inv.customer_id
    try:
        veh = db.session.query(Vehicle).filter(db.func.upper(Vehicle.vin) == vin_input).first()
        if veh:
            vehicle_id = veh.id
            # Prefer explicit owner as the customer; fallback to invoice customer
            if getattr(veh, 'owner_customer_id', None):
                customer_id = veh.owner_customer_id
    except Exception:
        pass
    p = Payment(invoice_id=inv.id, amount_omr=amt, method=method, reference=reference,
                vehicle_id=vehicle_id, customer_id=customer_id)
    db.session.add(p)
    # If this invoice looks like a car purchase (items linked to a vehicle) and has no explicit type,
    # classify it as a CAR invoice so dashboards and reports treat it correctly.
    try:
        if not (inv.invoice_type and str(inv.invoice_type).strip()):
            it = (
                db.session.query(InvoiceItem)
                .filter(InvoiceItem.invoice_id == inv.id, InvoiceItem.vehicle_id.isnot(None))
                .first()
            )
            if it:
                inv.invoice_type = 'CAR'
                if not getattr(inv, 'vehicle_id', None):
                    inv.vehicle_id = it.vehicle_id
    except Exception:
        # Non-blocking classification
        pass
    # update status
    paid = inv.paid_total() + amt
    if paid >= (inv.total_omr or 0):
        inv.status = 'Paid'
    else:
        inv.status = 'Partial'
    # If this is a service invoice (not CAR), prefer AR settlement if revenue already recognized
    if inv.invoice_type and inv.invoice_type != 'CAR' and float(amt) > 0:
        try:
            # Has revenue been recognized for this invoice? If yes: Dr Bank / Cr AR. Else: Dr Bank / Cr Revenue.
            recognized = (
                db.session.query(JournalEntry)
                .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
                .join(Account, JournalLine.account_id == Account.id)
                .filter(JournalEntry.invoice_id == inv.id, Account.code.like('R%'))
                .first()
            )
            if recognized:
                ar_code = _get_client_account_code(inv.customer_id, 'receivable', 'A300')
                _post_journal(
                    description='Service invoice payment', reference=inv.invoice_number,
                    lines=[('A100', float(amt), 0.0), (ar_code, 0.0, float(amt))],
                    customer_id=customer_id, vehicle_id=vehicle_id, invoice_id=inv.id, is_client_fund=False,
                )
            else:
                _post_journal(
                    description='Commission/service payment', reference=inv.invoice_number,
                    lines=[('A100', float(amt), 0.0), (_get_client_account_code(inv.customer_id, 'service', 'R300'), 0.0, float(amt))],
                    customer_id=customer_id, vehicle_id=vehicle_id, invoice_id=inv.id, is_client_fund=False,
                )
        except Exception:
            # Non-blocking
            pass
    try:
        db.session.commit()
        flash(_('Payment recorded'), 'success')
    except Exception:
        db.session.rollback()
        flash(_('Failed to save payment'), 'danger')
    return redirect(url_for('acct.payments_list'))


# Bill of Lading
@acct_bp.route('/bol')
@role_required('accountant', 'admin')
def bol_list():
    bols = db.session.query(BillOfLading).order_by(BillOfLading.created_at.desc()).all()
    shipments = db.session.query(Shipment).order_by(Shipment.created_at.desc()).all()
    return render_template('accounting/bol_list.html', bols=bols, shipments=shipments)


@acct_bp.route('/bol/new', methods=['POST'])
@role_required('accountant', 'admin')
def bol_new():
    shipment_id = request.form.get('shipment_id')
    bol_number = request.form.get('bol_number') or f"BOL-{int(datetime.utcnow().timestamp())}"
    bol = BillOfLading(bol_number=bol_number, shipment_id=int(shipment_id) if shipment_id else None)
    db.session.add(bol)
    try:
        db.session.commit()
        flash(_('BOL created'), 'success')
    except Exception:
        db.session.rollback()
        flash(_('Failed to create BOL'), 'danger')
    return redirect(url_for('acct.bol_list'))


@acct_bp.route('/bol/<int:bol_id>/export')
@role_required('accountant', 'admin')
def bol_export(bol_id: int):
    bol = db.session.get(BillOfLading, bol_id)
    if not bol:
        abort(404)
    vehicles = db.session.query(Vehicle).join(VehicleShipment, Vehicle.id == VehicleShipment.vehicle_id).\
        filter(VehicleShipment.shipment_id == bol.shipment_id).all()
    path = render_bol_pdf(bol, vehicles)
    bol.pdf_path = path
    db.session.commit()
    return send_file(path, as_attachment=True, download_name=f"{bol.bol_number}.pdf")


@acct_bp.route('/bol/<int:bol_id>/email', methods=['POST'])
@role_required('accountant', 'admin')
def bol_email(bol_id: int):
    recipient = (request.form.get('email') or '').strip()
    if not recipient:
        flash(_('Recipient email required'), 'danger'); return redirect(url_for('acct.bol_list'))
    bol = db.session.get(BillOfLading, bol_id)
    if not bol:
        flash(_('BOL not found'), 'danger'); return redirect(url_for('acct.bol_list'))
    vehicles = db.session.query(Vehicle).join(VehicleShipment, Vehicle.id == VehicleShipment.vehicle_id).\
        filter(VehicleShipment.shipment_id == bol.shipment_id).all()
    path = bol.pdf_path or render_bol_pdf(bol, vehicles)
    try:
        msg = Message(subject=_('BOL %(n)s', n=bol.bol_number), recipients=[recipient])
        msg.body = _('Please find attached Bill of Lading %(n)s.', n=bol.bol_number)
        with open(path, 'rb') as f:
            msg.attach(filename=f"{bol.bol_number}.pdf", content_type='application/pdf', data=f.read())
        mail.send(msg)
        flash(_('Email sent'), 'success')
    except Exception:
        flash(_('Failed to send email'), 'danger')
    return redirect(url_for('acct.bol_list'))


@acct_bp.route('/bol/<int:bol_id>/upload', methods=['POST'])
@role_required('accountant', 'admin')
def bol_upload(bol_id: int):
    bol = db.session.get(BillOfLading, bol_id)
    if not bol:
        flash(_('BOL not found'), 'danger'); return redirect(url_for('acct.bol_list'))
    f = request.files.get('file')
    if not f:
        flash(_('No file uploaded'), 'danger'); return redirect(url_for('acct.bol_list'))
    import os
    outdir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'bols')
    os.makedirs(outdir, exist_ok=True)
    filename = f"{bol.bol_number}.pdf"
    path = os.path.join(outdir, filename)
    f.save(path)
    bol.pdf_path = path
    db.session.commit()
    flash(_('BOL uploaded'), 'success')
    return redirect(url_for('acct.bol_list'))


# Reports
@acct_bp.route('/reports')
@role_required('accountant', 'admin')
def reports():
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from openpyxl import Workbook
    report_type = request.args.get('type', 'monthly')
    export = request.args.get('export')

    now = datetime.utcnow()
    if report_type == 'monthly':
        labels, revenue, expenses = [], [], []
        dt = datetime(now.year, now.month, 1)
        usd_to_omr = float(current_app.config.get('OMR_EXCHANGE_RATE', 0.385))
        for month_index in range(12):
            start = dt
            end = datetime(dt.year + 1, 1, 1) if dt.month == 12 else datetime(dt.year, dt.month + 1, 1)
            labels.append(dt.strftime('%b %Y'))
            # IFRS: revenue from GL R* excluding client funds
            rev = (
                db.session.query(db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0))
                .join(Account, JournalLine.account_id == Account.id)
                .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
                .filter(
                    JournalEntry.entry_date >= start,
                    JournalEntry.entry_date < end,
                    Account.code.like('R%'),
                    JournalEntry.is_client_fund.is_(False),
                )
                .scalar()
                or 0
            )
            # Expenses: combine operational costs + car purchase totals in period
            car_cost = db.session.query(db.func.coalesce(db.func.sum(Invoice.total_omr), 0)).\
                filter(Invoice.created_at >= start, Invoice.created_at < end, Invoice.status == 'Paid', Invoice.invoice_type == 'CAR').scalar() or 0
            freight = db.session.query(db.func.coalesce(db.func.sum(Shipment.cost_freight_usd), 0)).filter(Shipment.created_at >= start, Shipment.created_at < end).scalar() or 0
            customs = db.session.query(db.func.coalesce(db.func.sum(InternationalCost.customs_omr), 0)).filter(InternationalCost.created_at >= start, InternationalCost.created_at < end).scalar() or 0
            vat = db.session.query(db.func.coalesce(db.func.sum(InternationalCost.vat_omr), 0)).filter(InternationalCost.created_at >= start, InternationalCost.created_at < end).scalar() or 0
            local_t = db.session.query(db.func.coalesce(db.func.sum(InternationalCost.local_transport_omr), 0)).filter(InternationalCost.created_at >= start, InternationalCost.created_at < end).scalar() or 0
            misc = db.session.query(db.func.coalesce(db.func.sum(InternationalCost.misc_omr), 0)).filter(InternationalCost.created_at >= start, InternationalCost.created_at < end).scalar() or 0
            exp = float(freight) * usd_to_omr + float(customs or 0) + float(vat or 0) + float(local_t or 0) + float(misc or 0) + float(car_cost or 0)
            revenue.append(float(rev)); expenses.append(float(exp))
            if dt.month == 1: dt = datetime(dt.year - 1, 12, 1)
            else: dt = datetime(dt.year, dt.month - 1, 1)
        labels, revenue, expenses = list(reversed(labels)), list(reversed(revenue)), list(reversed(expenses))

        if export == 'pdf':
            buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4)
            width, height = A4; y = height - 40
            c.setFont('Helvetica-Bold', 16); c.drawString(40, y, _('Monthly Profit & Loss'))
            y -= 25; c.setFont('Helvetica-Bold', 11);
            c.drawString(40, y, _('Month')); c.drawString(200, y, _('Revenue')); c.drawString(320, y, _('Expenses')); c.drawString(440, y, _('Profit'))
            y -= 14; c.setFont('Helvetica', 10)
            for m, r, e in zip(labels, revenue, expenses):
                if y < 40:
                    c.showPage(); y = height - 40; c.setFont('Helvetica-Bold', 11)
                    c.drawString(40, y, _('Month')); c.drawString(200, y, _('Revenue')); c.drawString(320, y, _('Expenses')); c.drawString(440, y, _('Profit'))
                    y -= 14; c.setFont('Helvetica', 10)
                c.drawString(40, y, m); c.drawRightString(300, y, f"{r:,.3f}"); c.drawRightString(420, y, f"{e:,.3f}"); c.drawRightString(540, y, f"{(r-e):,.3f}"); y -= 12
            c.showPage(); c.save(); buf.seek(0)
            return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name='monthly_pl.pdf')
        
        
    if report_type == 'monthly' and export == 'xlsx':
        wb = Workbook(); ws = wb.active; ws.title = 'Monthly P&L'; ws.append([_('Month'),_('Revenue'),_('Expenses'),_('Profit')])
        for m, r, e in zip(labels, revenue, expenses): ws.append([m, r, e, r-e])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='monthly_pl.xlsx')

    if report_type == 'by_client':
        # Aggregate invoices by client
        # Sum GL revenue by customer (R* accounts) excluding client funds
        rows = (
            db.session.query(Customer.company_name, db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0))
            .join(JournalEntry, JournalEntry.customer_id == Customer.id, isouter=True)
            .join(JournalLine, JournalLine.entry_id == JournalEntry.id, isouter=True)
            .join(Account, JournalLine.account_id == Account.id, isouter=True)
            .filter(Account.code.like('R%'), JournalEntry.is_client_fund.is_(False))
            .group_by(Customer.company_name)
            .order_by(Customer.company_name.asc())
            .all()
        )
        data = [(name or '-', float(total or 0)) for name, total in rows]
        headers = [_('Client'), _('Total (OMR)')]
        if export == 'xlsx':
            wb = Workbook(); ws = wb.active; ws.title = 'Invoices by Client'; ws.append([_('Client'), _('Total (OMR)')])
            for n, t in data: ws.append([n, t])
            buf = BytesIO(); wb.save(buf); buf.seek(0)
            return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='invoices_by_client.xlsx')
        if export == 'pdf':
            buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4)
            width, height = A4; y = height - 40
            c.setFont('Helvetica-Bold', 16); c.drawString(40, y, _('Invoices by Client')); y -= 20; c.setFont('Helvetica', 10)
            for n, t in data:
                if y < 40: c.showPage(); y = height - 40; c.setFont('Helvetica', 10)
                c.drawString(40, y, n); c.drawRightString(550, y, f"{t:,.3f}"); y -= 14
            c.showPage(); c.save(); buf.seek(0)
            return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name='invoices_by_client.pdf')
        return render_template('accounting/reports.html', report_type='by_client', table=data, headers=headers)

    if report_type == 'taxes':
        # Monthly customs and VAT
        labels, customs_m, vat_m = [], [], []
        dt = datetime(now.year, now.month, 1)
        for month_index in range(12):
            start = dt
            end = datetime(dt.year + 1, 1, 1) if dt.month == 12 else datetime(dt.year, dt.month + 1, 1)
            labels.append(dt.strftime('%b %Y'))
            customs = db.session.query(db.func.coalesce(db.func.sum(InternationalCost.customs_omr), 0)).\
                filter(InternationalCost.created_at >= start, InternationalCost.created_at < end).scalar() or 0
            vat = db.session.query(db.func.coalesce(db.func.sum(InternationalCost.vat_omr), 0)).\
                filter(InternationalCost.created_at >= start, InternationalCost.created_at < end).scalar() or 0
            customs_m.append(float(customs)); vat_m.append(float(vat))
            if dt.month == 1: dt = datetime(dt.year - 1, 12, 1)
            else: dt = datetime(dt.year, dt.month - 1, 1)
        labels, customs_m, vat_m = list(reversed(labels)), list(reversed(customs_m)), list(reversed(vat_m))
        if export == 'xlsx':
            wb = Workbook(); ws = wb.active; ws.title = 'Taxes'; ws.append([_('Month'), _('Customs (OMR)'), _('VAT (OMR)')])
            for m, cst, vt in zip(labels, customs_m, vat_m): ws.append([m, cst, vt])
            buf = BytesIO(); wb.save(buf); buf.seek(0)
            return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='taxes.xlsx')
        if export == 'pdf':
            buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4)
            width, height = A4; y = height - 40
            c.setFont('Helvetica-Bold', 16); c.drawString(40, y, _('Customs & VAT by Month')); y -= 20; c.setFont('Helvetica', 10)
            for m, cst, vt in zip(labels, customs_m, vat_m):
                if y < 40: c.showPage(); y = height - 40; c.setFont('Helvetica', 10)
                c.drawString(40, y, m); c.drawRightString(320, y, f"{cst:,.3f}"); c.drawRightString(560, y, f"{vt:,.3f}"); y -= 14
            c.showPage(); c.save(); buf.seek(0)
            return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name='taxes.pdf')
        return render_template('accounting/reports.html', report_type='taxes', chart={"months": labels, "customs": customs_m, "vat": vat_m})

    if report_type == 'balance_sheet':
        # Balance Sheet with Client Deposits under Current Liabilities
        def sum_acct(prefix: str, exclude_client_fund: bool | None = None):
            from ...models import Account, JournalLine, JournalEntry
            q = db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0)).\
                join(Account, JournalLine.account_id == Account.id).\
                join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
                filter(Account.code.like(f"{prefix}%"))
            if exclude_client_fund is True:
                q = q.filter(JournalEntry.is_client_fund.is_(False))
            elif exclude_client_fund is False:
                q = q.filter(JournalEntry.is_client_fund.is_(True))
            return float(q.scalar() or 0)
        assets = sum_acct('A', exclude_client_fund=True)
        # Total liabilities including client funds
        liabilities_total = -sum_acct('L', exclude_client_fund=None)
        # Client deposits account (L200*) balance from all entries (client funds scope only preferred)
        client_deposits = -(
            db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0))
            .join(Account, JournalLine.account_id == Account.id)
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(Account.code.like('L200%'))
            .scalar() or 0
        )
        other_liabilities = max(0.0, liabilities_total - client_deposits)
        equity = assets - (client_deposits + other_liabilities)
        data = [
            (_('Assets'), assets),
            (_('Client Deposits (Current Liabilities)'), client_deposits),
            (_('Other Liabilities'), other_liabilities),
            (_('Equity'), equity),
        ]
        headers = [_('Category'), _('Amount (OMR)')]
        return render_template('accounting/reports.html', report_type='balance_sheet', table=data, headers=headers)

    if report_type == 'trial_balance':
        # Trial balance from GL (all entries)
        rows = (
            db.session.query(
                Account.code,
                Account.name,
                db.func.coalesce(db.func.sum(JournalLine.debit), 0),
                db.func.coalesce(db.func.sum(JournalLine.credit), 0),
            )
            .join(JournalLine, JournalLine.account_id == Account.id)
            .group_by(Account.code, Account.name)
            .order_by(Account.code.asc())
            .all()
        )
        data = [(code, name, float(dr or 0), float(cr or 0), float((dr or 0) - (cr or 0))) for code, name, dr, cr in rows]
        headers = [_('Account Code'), _('Account Name'), _('Debit'), _('Credit'), _('Net (Dr-Cr)')]
        return render_template('accounting/reports.html', report_type='trial_balance', table=data, headers=headers)

    if report_type == 'general_ledger':
        # General ledger for a specific account
        acct_code = (request.args.get('code') or 'A100').strip()
        acct = db.session.query(Account).filter(Account.code == acct_code).first()
        if not acct:
            return render_template('accounting/reports.html', report_type='general_ledger', table=[], headers=[_('Date'), _('Description'), _('Debit'), _('Credit')])
        rows = (
            db.session.query(JournalEntry.entry_date, JournalEntry.description, JournalLine.debit, JournalLine.credit)
            .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
            .filter(JournalLine.account_id == acct.id)
            .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc())
            .limit(1000)
            .all()
        )
        data = [(
            (dt.strftime('%Y-%m-%d') if dt else ''), desc or '-', float(dr or 0), float(cr or 0)
        ) for dt, desc, dr, cr in rows]
        headers = [_('Date'), _('Description'), _('Debit'), _('Credit')]
        return render_template('accounting/reports.html', report_type='general_ledger', table=data, headers=headers)

    if report_type == 'cash_flow':
        # Simple cash flow (Direct): monthly net cash movement on Bank accounts (A100*)
        method = (request.args.get('method') or 'direct').strip().lower()
        now = datetime.utcnow()
        dt = datetime(now.year, now.month, 1)
        labels = []
        net = []
        for month_index in range(12):
            start = dt
            end = datetime(dt.year + 1, 1, 1) if dt.month == 12 else datetime(dt.year, dt.month + 1, 1)
            labels.append(dt.strftime('%b %Y'))
            q = (
                db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0))
                .join(Account, JournalLine.account_id == Account.id)
                .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
                .filter(
                    JournalEntry.entry_date >= start,
                    JournalEntry.entry_date < end,
                    Account.code.like('A100%'),
                    JournalEntry.is_client_fund.is_(False),
                )
            )
            val = float(q.scalar() or 0)
            net.append(val)
            if dt.month == 1: dt = datetime(dt.year - 1, 12, 1)
            else: dt = datetime(dt.year, dt.month - 1, 1)
        labels, net = list(reversed(labels)), list(reversed(net))
        headers = [_('Month'), _('Net Cash Movement (OMR)')]
        table = list(zip(labels, net))
        return render_template('accounting/reports.html', report_type='cash_flow', table=table, headers=headers)

    if report_type == 'ar_aging':
        # Accounts receivable by customer from invoices minus payments
        rows = db.session.query(Customer.company_name, db.func.coalesce(db.func.sum(Invoice.total_omr), 0) - db.func.coalesce(db.func.sum(Payment.amount_omr), 0)).\
            join(Invoice, Invoice.customer_id == Customer.id, isouter=True).\
            join(Payment, Payment.invoice_id == Invoice.id, isouter=True).\
            group_by(Customer.company_name).all()
        data = [(n or '-', float(bal or 0)) for n, bal in rows]
        headers = [_('Client'), _('Balance (OMR)')]
        return render_template('accounting/reports.html', report_type='ar_aging', table=data, headers=headers)

    if report_type == 'inventory_by_vehicle':
        # Inventory value per vehicle (capitalized purchase OMR)
        from ...models import Vehicle
        rate = Decimal(str(current_app.config.get('OMR_EXCHANGE_RATE', 0.385)))
        rows = db.session.query(Vehicle.vin, Vehicle.make, Vehicle.model, Vehicle.year, Vehicle.purchase_price_usd).all()
        data = [(vin, make, model, year, float((Decimal(str(pp or 0)) * rate))) for vin, make, model, year, pp in rows]
        headers = ['VIN', _('Make'), _('Model'), _('Year'), _('Value (OMR)')]
        return render_template('accounting/reports.html', report_type='inventory_by_vehicle', table=data, headers=headers)

    if report_type == 'fines_revenue':
        # Sum fines revenue (R300), excluding client fund flagged entries
        total = db.session.query(db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0)).\
            join(Account, JournalLine.account_id == Account.id).\
            join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
            filter(Account.code == 'R300', JournalEntry.is_client_fund.is_(False)).scalar() or 0
        headers = [_('Metric'), _('Amount (OMR)')]
        return render_template('accounting/reports.html', report_type='fines_revenue', table=[[str(_('Fines Revenue')), float(total)]], headers=headers)

    if report_type == 'customer_statement':
        # Detailed statement for a single customer
        try:
            customer_id = int(request.args.get('customer_id'))
        except Exception:
            customer_id = None
        if not customer_id:
            return render_template('accounting/reports.html', report_type='customer_statement', table=[], headers=[_('Date'), _('Description'), _('Debit'), _('Credit'), _('Balance')])
        rows = []
        balance = Decimal('0')
        invs = db.session.query(Invoice).filter(Invoice.customer_id == customer_id).order_by(Invoice.created_at.asc()).all()
        for inv in invs:
            amt = Decimal(str(inv.total_omr or 0))
            balance += amt
            rows.append([inv.created_at.strftime('%Y-%m-%d') if inv.created_at else '', f"Invoice {inv.invoice_number}", float(amt), 0.0, float(balance)])
            pays = db.session.query(Payment).filter(Payment.invoice_id == inv.id).order_by(Payment.received_at.asc()).all()
            for p in pays:
                val = Decimal(str(p.amount_omr or 0))
                balance -= val
                rows.append([p.received_at.strftime('%Y-%m-%d') if p.received_at else '', f"Payment {p.reference or ''}", 0.0, float(val), float(balance)])
        headers = [_('Date'), _('Description'), _('Debit'), _('Credit'), _('Balance')]
        return render_template('accounting/reports.html', report_type='customer_statement', table=rows, headers=headers)

    # default monthly P&L chart
    return render_template('accounting/reports.html', report_type='monthly', chart={"months": labels, "revenue": revenue, "expenses": expenses})


# Accounting Settings (Accountant scope)
@acct_bp.route('/settings', methods=['GET','POST'])
@role_required('accountant', 'admin')
def settings():
    settings_row = db.session.query(Setting).first()
    if not settings_row:
        settings_row = Setting(customs_rate=0, vat_rate=0, shipping_fee=0, insurance_rate=0)
        db.session.add(settings_row)
        db.session.commit()
    if request.method == 'POST':
        try:
            settings_row.customs_rate = float(request.form.get('customs_rate') or 0)
            settings_row.vat_rate = float(request.form.get('vat_rate') or 0)
            settings_row.shipping_fee = float(request.form.get('shipping_fee') or 0)
            settings_row.insurance_rate = float(request.form.get('insurance_rate') or 0)
            db.session.commit(); flash('Settings saved', 'success')
        except Exception:
            db.session.rollback(); flash('Failed to save', 'danger')
    return render_template('accounting/settings.html', settings=settings_row)


# Chart of Accounts (simple list and create)
@acct_bp.route('/accounts')
@role_required('accountant', 'admin')
def accounts_list():
    rows = db.session.query(Account).order_by(Account.code.asc()).all()
    return render_template('accounting/accounts_list.html', accounts=rows)


# ---- Vehicle Statement of Account (SOA) ----
@acct_bp.route('/vehicles')
@role_required('accountant', 'admin')
def vehicles_list():
    q = db.session.query(Vehicle).order_by(Vehicle.created_at.desc()).limit(200)
    client_id = request.args.get('client_id')
    if client_id:
        try:
            q = q.filter(Vehicle.owner_customer_id == int(client_id))
        except Exception:
            pass
    vehicles = q.all()
    # Precompute running balances from journal per vehicle
    balances = {}
    for v in vehicles:
        total = db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0)).\
            join(Account, JournalLine.account_id == Account.id).\
            join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
            filter(JournalEntry.vehicle_id == v.id).scalar() or 0
        balances[v.id] = float(total)
    return render_template('accounting/vehicles_list.html', vehicles=vehicles, balances=balances)


@acct_bp.route('/vehicles/<int:vehicle_id>/statement')
@role_required('accountant', 'admin')
def vehicle_statement(vehicle_id: int):
    v = db.session.get(Vehicle, vehicle_id)
    if not v:
        abort(404)
    # Fetch journal lines for this vehicle
    rows = (
        db.session.query(JournalEntry.entry_date, JournalEntry.description, Account.code, Account.name, JournalLine.debit, JournalLine.credit)
        .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
        .join(Account, JournalLine.account_id == Account.id)
        .filter(JournalEntry.vehicle_id == vehicle_id)
        .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc(), JournalLine.id.asc())
        .all()
    )
    # Build running balance
    statement = []
    running = 0.0
    for dt, desc, code, name, dr, cr in rows:
        dr_f = float(dr or 0)
        cr_f = float(cr or 0)
        running += dr_f - cr_f
        statement.append({
            'date': dt.strftime('%Y-%m-%d') if dt else '',
            'description': desc,
            'account_code': code,
            'account_name': name,
            'debit': dr_f,
            'credit': cr_f,
            'balance': running,
        })
    # Totals breakdown by vehicle-level semantic kinds
    def sum_kind(kind: str, default_prefix: str) -> float:
        codes = []
        vas = db.session.query(VehicleAccountStructure).filter_by(vehicle_id=vehicle_id).first()
        if vas:
            code_val = {
                'auction': vas.auction_account_code,
                'freight': vas.freight_account_code,
                'customs': vas.customs_account_code,
                'commission': vas.commission_account_code,
                'storage': vas.storage_account_code,
                'deposit': vas.deposit_account_code,
            }.get(kind)
            if code_val:
                codes.append(code_val)
        q = db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0)).\
            join(Account, JournalLine.account_id == Account.id).\
            join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
            filter(JournalEntry.vehicle_id == vehicle_id)
        if codes:
            q = q.filter(Account.code.in_(codes))
        else:
            q = q.filter(Account.code.like(f"{default_prefix}%"))
        return float(q.scalar() or 0)

    totals = {
        'auction_cost_omr': sum_kind('auction', 'A150'),
        'freight_omr': sum_kind('freight', 'E200'),
        'customs_omr': sum_kind('customs', 'E220'),
        'service_fee_omr': -sum_kind('commission', 'R300'),  # credit balances as positive
        'deposit_net_omr': sum_kind('deposit', 'L200'),
    }
    totals['outstanding_balance_omr'] = statement[-1]['balance'] if statement else 0.0

    return render_template('accounting/vehicle_statement.html', vehicle=v, statement=statement, totals=totals)


@acct_bp.route('/vehicles/<int:vehicle_id>/payments', methods=['POST'])
@role_required('accountant', 'admin')
def vehicle_statement_add_payment(vehicle_id: int):
    """Record a payment (client fund deposit) for this vehicle with custom description.

    Posts: Dr Bank (A100) / Cr Vehicle Deposit (L200-V{vehicle_id}) as client funds.
    The user-provided description is stored on the journal so it appears in the statement.
    """
    v = db.session.get(Vehicle, vehicle_id)
    if not v:
        abort(404)
    # Inputs
    amount_val = _parse_number_input(request.form.get('amount'))
    description = (request.form.get('description') or '').strip()
    method = (request.form.get('method') or '').strip() or None

    if amount_val <= 0:
        flash(_('Invalid amount'), 'danger')
        return redirect(url_for('acct.vehicle_statement', vehicle_id=vehicle_id))

    # Prefer vehicle owner as customer; fallback to None
    customer_id = getattr(v, 'owner_customer_id', None)

    # Create a CustomerDeposit record for traceability
    dep = CustomerDeposit(
        customer_id=customer_id,
        vehicle_id=vehicle_id,
        auction_id=getattr(v, 'auction_id', None),
        amount_omr=float(amount_val),
        method=method,
        reference=description or None,
        status='held',
    )
    db.session.add(dep)

    # Journal: Dr Bank / Cr Vehicle Deposit (client fund)
    dep_code = _get_vehicle_account_code(vehicle_id, 'deposit', _get_client_account_code(customer_id, 'deposit', 'L200'))
    _post_journal(
        description=description or _('Vehicle payment received'),
        reference=(v.vin or str(vehicle_id)),
        lines=[('A100', float(amount_val), 0.0), (dep_code, 0.0, float(amount_val))],
        customer_id=customer_id,
        vehicle_id=vehicle_id,
        auction_id=getattr(v, 'auction_id', None),
        is_client_fund=True,
    )

    try:
        db.session.commit()
        flash(_('Payment recorded'), 'success')
    except Exception:
        db.session.rollback()
        flash(_('Failed to save payment'), 'danger')
    return redirect(url_for('acct.vehicle_statement', vehicle_id=vehicle_id))


@acct_bp.route('/vehicles/<int:vehicle_id>/statement.pdf')
@role_required('accountant', 'admin')
def vehicle_statement_pdf(vehicle_id: int):
    v = db.session.get(Vehicle, vehicle_id)
    if not v:
        abort(404)
    # Recompute same data as HTML view
    rows = (
        db.session.query(JournalEntry.entry_date, JournalEntry.description, Account.code, Account.name, JournalLine.debit, JournalLine.credit)
        .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
        .join(Account, JournalLine.account_id == Account.id)
        .filter(JournalEntry.vehicle_id == vehicle_id)
        .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc(), JournalLine.id.asc())
        .all()
    )
    statement = []
    running = 0.0
    for dt, desc, code, name, dr, cr in rows:
        dr_f = float(dr or 0)
        cr_f = float(cr or 0)
        running += dr_f - cr_f
        statement.append({
            'date': dt.strftime('%Y-%m-%d') if dt else '',
            'description': desc,
            'account_code': code,
            'account_name': name,
            'debit': dr_f,
            'credit': cr_f,
            'balance': running,
        })
    def sum_kind(kind: str, default_prefix: str) -> float:
        codes = []
        vas = db.session.query(VehicleAccountStructure).filter_by(vehicle_id=vehicle_id).first()
        if vas:
            code_val = {
                'auction': vas.auction_account_code,
                'freight': vas.freight_account_code,
                'customs': vas.customs_account_code,
                'commission': vas.commission_account_code,
                'storage': vas.storage_account_code,
                'deposit': vas.deposit_account_code,
            }.get(kind)
            if code_val:
                codes.append(code_val)
        q = db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0)).\
            join(Account, JournalLine.account_id == Account.id).\
            join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
            filter(JournalEntry.vehicle_id == vehicle_id)
        if codes:
            q = q.filter(Account.code.in_(codes))
        else:
            q = q.filter(Account.code.like(f"{default_prefix}%"))
        return float(q.scalar() or 0)
    totals = {
        'auction_cost_omr': sum_kind('auction', 'A150'),
        'freight_omr': sum_kind('freight', 'E200'),
        'customs_omr': sum_kind('customs', 'E220'),
        'service_fee_omr': -sum_kind('commission', 'R300'),
        'deposit_net_omr': sum_kind('deposit', 'L200'),
    }
    totals['outstanding_balance_omr'] = statement[-1]['balance'] if statement else 0.0
    path = render_vehicle_statement_pdf(v, statement, totals)
    return send_file(path, as_attachment=True, download_name=f"vehicle_statement_{v.vin or v.id}.pdf")


# ---- API Endpoints ----
@acct_bp.get('/api/vehicles/<int:vehicle_id>/statement')
@login_required
def api_vehicle_statement(vehicle_id: int):
    # limit to staff or vehicle owner
    v = db.session.get(Vehicle, vehicle_id)
    if not v:
        return jsonify({'error': 'not found'}), 404
    try:
        from flask_login import current_user
        is_staff = bool(getattr(current_user, 'role', None) and getattr(current_user.role, 'name', '').lower() in {'admin','employee','accountant'})
        is_owner = False
        if getattr(current_user, 'id', None):
            cust = db.session.query(Customer).filter(Customer.user_id == current_user.id).first()
            is_owner = bool(cust and v.owner_customer_id == cust.id)
        if not (is_staff or is_owner):
            return jsonify({'error': 'forbidden'}), 403
    except Exception:
        return jsonify({'error': 'forbidden'}), 403

    rows = (
        db.session.query(JournalEntry.entry_date, JournalEntry.description, Account.code, Account.name, JournalLine.debit, JournalLine.credit)
        .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
        .join(Account, JournalLine.account_id == Account.id)
        .filter(JournalEntry.vehicle_id == vehicle_id)
        .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc(), JournalLine.id.asc())
        .all()
    )
    data = []
    running = 0.0
    for dt, desc, code, name, dr, cr in rows:
        dr_f = float(dr or 0)
        cr_f = float(cr or 0)
        running += dr_f - cr_f
        data.append({
            'date': dt.strftime('%Y-%m-%d') if dt else '',
            'description': desc,
            'account_code': code,
            'account_name': name,
            'debit': dr_f,
            'credit': cr_f,
            'balance': running,
        })
    # Totals using same logic as HTML/PDF (respect vehicle-specific account mapping)
    def sum_kind(kind: str, default_prefix: str) -> float:
        codes = []
        vas = db.session.query(VehicleAccountStructure).filter_by(vehicle_id=vehicle_id).first()
        if vas:
            code_val = {
                'auction': vas.auction_account_code,
                'freight': vas.freight_account_code,
                'customs': vas.customs_account_code,
                'commission': vas.commission_account_code,
                'storage': vas.storage_account_code,
                'deposit': vas.deposit_account_code,
            }.get(kind)
            if code_val:
                codes.append(code_val)
        q = (
            db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0))
            .join(Account, JournalLine.account_id == Account.id)
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(JournalEntry.vehicle_id == vehicle_id)
        )
        if codes:
            q = q.filter(Account.code.in_(codes))
        else:
            q = q.filter(Account.code.like(f"{default_prefix}%"))
        return float(q.scalar() or 0)

    totals = {
        'auction_cost_omr': sum_kind('auction', 'A150'),
        'freight_omr': sum_kind('freight', 'E200'),
        'customs_omr': sum_kind('customs', 'E220'),
        'service_fee_omr': -sum_kind('commission', 'R300'),  # credit balances as positive
        'deposit_net_omr': sum_kind('deposit', 'L200'),
        'outstanding_balance_omr': data[-1]['balance'] if data else 0.0,
    }

    return jsonify({
        'vehicle_id': vehicle_id,
        'vin': v.vin,
        'client_id': v.owner_customer_id,
        'entries': data,
        'totals': totals,
    })


@acct_bp.get('/api/clients/<int:client_id>/vehicles/summary')
@login_required
def api_client_vehicles_summary(client_id: int):
    # staff or the client itself
    try:
        from flask_login import current_user
        is_staff = bool(getattr(current_user, 'role', None) and getattr(current_user.role, 'name', '').lower() in {'admin','employee','accountant'})
        is_self = False
        if getattr(current_user, 'id', None):
            cust = db.session.query(Customer).filter(Customer.user_id == current_user.id).first()
            is_self = bool(cust and cust.id == client_id)
        if not (is_staff or is_self):
            return jsonify({'error': 'forbidden'}), 403
    except Exception:
        return jsonify({'error': 'forbidden'}), 403

    vehicles = db.session.query(Vehicle).filter(Vehicle.owner_customer_id == client_id).all()
    out = []
    for v in vehicles:
        total = db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0)).\
            join(Account, JournalLine.account_id == Account.id).\
            join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
            filter(JournalEntry.vehicle_id == v.id).scalar() or 0
        out.append({
            'vehicle_id': v.id,
            'vin': v.vin,
            'status': v.status,
            'balance_omr': float(total),
        })
    return jsonify({'client_id': client_id, 'vehicles': out})


@acct_bp.route('/accounts/new', methods=['GET','POST'])
@role_required('accountant', 'admin')
def accounts_new():
    if request.method == 'POST':
        code = (request.form.get('code') or '').strip()
        name = (request.form.get('name') or '').strip()
        typ = (request.form.get('type') or '').strip().upper()
        if not code or not name or typ not in {'ASSET','LIABILITY','EQUITY','REVENUE','EXPENSE'}:
            flash(_('Please fill in all fields.'), 'danger')
            return render_template('accounting/accounts_form.html')
        if db.session.query(Account).filter(Account.code == code).first():
            flash(_('Account code already exists.'), 'danger')
            return render_template('accounting/accounts_form.html')
        db.session.add(Account(code=code, name=name, type=typ))
        try:
            db.session.commit(); flash(_('Account created'), 'success')
            return redirect(url_for('acct.accounts_list'))
        except Exception:
            db.session.rollback(); flash(_('Failed to create account'), 'danger')
    return render_template('accounting/accounts_form.html')


# ---- Client Accounting View ----
@acct_bp.route('/clients/view')
@role_required('accountant', 'admin')
def client_view():
    try:
        customer_id = int(request.args.get('customer_id')) if request.args.get('customer_id') else None
    except Exception:
        customer_id = None
    customers = db.session.query(Customer).order_by(Customer.company_name.asc()).all()
    customer = db.session.get(Customer, customer_id) if customer_id else None
    ledger = []
    deposits = []
    auction_ledger = []
    service_rows = []
    balances = {"deposits": 0.0, "ar": 0.0, "paid": 0.0, "revenue": 0.0}
    pl = {"revenue": 0.0, "logistics": 0.0}
    if customer:
        # Ensure sub-accounts exist for this client
        try:
            _ensure_client_accounts(customer)
        except Exception:
            pass
        cas = db.session.query(ClientAccountStructure).filter_by(customer_id=customer.id).first()
        dep_code = cas.deposit_account_code if cas else 'L200'
        srv_code = cas.service_revenue_account_code if cas else 'R300'
        log_code = cas.logistics_expense_account_code if cas else 'E200'

        # Ledger: all entries for this client
        rows = (
            db.session.query(JournalEntry.entry_date, JournalEntry.description, JournalLine.debit, JournalLine.credit)
            .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
            .filter(JournalEntry.customer_id == customer.id)
            .order_by(JournalEntry.entry_date.asc(), JournalEntry.id.asc())
            .limit(1000)
            .all()
        )
        ledger = [{
            'date': (dt.strftime('%Y-%m-%d') if dt else ''),
            'desc': (desc or '-'),
            'debit': float(dr or 0),
            'credit': float(cr or 0),
        } for dt, desc, dr, cr in rows]

        # Deposits and refunds
        dep_rows = db.session.query(CustomerDeposit).filter(CustomerDeposit.customer_id == customer.id).order_by(CustomerDeposit.created_at.asc()).all()
        for d in dep_rows:
            deposits.append({
                'date': d.received_at.strftime('%Y-%m-%d') if d.received_at else '',
                'reference': d.reference or '',
                'amount': float(d.amount_omr or 0),
                'type': d.status,
            })

        # Auction payments from client funds (client-fund journals with credit to Bank)
        auc_rows = (
            db.session.query(JournalEntry.entry_date, JournalEntry.description, JournalLine.debit, JournalLine.credit)
            .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
            .join(Account, JournalLine.account_id == Account.id)
            .filter(JournalEntry.customer_id == customer.id, JournalEntry.is_client_fund.is_(True), Account.code.like('A100%'))
            .order_by(JournalEntry.entry_date.asc())
            .all()
        )
        auction_ledger = [{
            'date': (dt.strftime('%Y-%m-%d') if dt else ''),
            'desc': desc or '-',
            'amount': float(cr or 0),
        } for dt, desc, dr, cr in auc_rows]

        # Service revenue rows for this client (non client-fund revenue)
        srv_rows = (
            db.session.query(JournalEntry.entry_date, JournalEntry.description, JournalLine.debit, JournalLine.credit)
            .join(JournalLine, JournalLine.entry_id == JournalEntry.id)
            .join(Account, JournalLine.account_id == Account.id)
            .filter(JournalEntry.customer_id == customer.id, JournalEntry.is_client_fund.is_(False), Account.code.like('R%'))
            .order_by(JournalEntry.entry_date.asc())
            .all()
        )
        service_rows = [{
            'date': (dt.strftime('%Y-%m-%d') if dt else ''),
            'desc': desc or '-',
            'amount': float(cr or 0) - float(dr or 0),
        } for dt, desc, dr, cr in srv_rows]

        # Balances
        balances['deposits'] = float(
            db.session.query(db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0))
            .join(Account, JournalLine.account_id == Account.id)
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(JournalEntry.customer_id == customer.id, Account.code == dep_code)
            .scalar() or 0
        )
        # Receivables: sum of AR account for this client
        ar_code = cas.receivable_account_code if cas else 'A300'
        balances['ar'] = float(
            db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0))
            .join(Account, JournalLine.account_id == Account.id)
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(JournalEntry.customer_id == customer.id, Account.code == ar_code)
            .scalar() or 0
        )
        # Paid total: net cash movements for this client (A100) excluding client fund
        balances['paid'] = float(
            db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0))
            .join(Account, JournalLine.account_id == Account.id)
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(JournalEntry.customer_id == customer.id, Account.code.like('A100%'), JournalEntry.is_client_fund.is_(False))
            .scalar() or 0
        )
        # Commission earned: revenue for this client excluding client fund
        balances['revenue'] = float(
            db.session.query(db.func.coalesce(db.func.sum(JournalLine.credit - JournalLine.debit), 0))
            .join(Account, JournalLine.account_id == Account.id)
            .join(JournalEntry, JournalLine.entry_id == JournalEntry.id)
            .filter(JournalEntry.customer_id == customer.id, Account.code.like('R%'), JournalEntry.is_client_fund.is_(False))
            .scalar() or 0
        )
        # Mini P&L
        pl['revenue'] = balances['revenue']
        # Logistics expense: sum of client's E200-Cxxxxx and E210-Cxxxxx if used
        log_total = db.session.query(db.func.coalesce(db.func.sum(JournalLine.debit - JournalLine.credit), 0)).\
            join(Account, JournalLine.account_id == Account.id).\
            join(JournalEntry, JournalLine.entry_id == JournalEntry.id).\
            filter(JournalEntry.customer_id == customer.id, db.or_(Account.code == log_code, Account.code == (log_code.replace('E200', 'E210'))), JournalEntry.is_client_fund.is_(False)).\
            scalar() or 0
        pl['logistics'] = float(log_total)

    export = (request.args.get('export') or '').strip().lower()
    if export in {'pdf','xlsx'} and customer:
        # Reuse existing reports export styles (customer_statement) if needed; for brevity, export the ledger
        if export == 'xlsx':
            from openpyxl import Workbook
            from io import BytesIO
            wb = Workbook(); ws = wb.active; ws.title = 'Client Statement'
            ws.append(['Date','Description','Debit','Credit'])
            for row in ledger:
                ws.append([row['date'], row['desc'], row['debit'], row['credit']])
            buf = BytesIO(); wb.save(buf); buf.seek(0)
            return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='client_statement.xlsx')
        if export == 'pdf':
            from io import BytesIO
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            buf = BytesIO(); c = canvas.Canvas(buf, pagesize=A4)
            width, height = A4; y = height - 40
            c.setFont('Helvetica-Bold', 16); c.drawString(40, y, f"Client Statement - {customer.display_name}"); y -= 20
            c.setFont('Helvetica', 10)
            for row in ledger:
                if y < 40:
                    c.showPage(); y = height - 40; c.setFont('Helvetica', 10)
                c.drawString(40, y, row['date']); c.drawString(120, y, row['desc'][:60]); c.drawRightString(450, y, f"{row['debit']:.3f}"); c.drawRightString(550, y, f"{row['credit']:.3f}"); y -= 12
            c.showPage(); c.save(); buf.seek(0)
            return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name='client_statement.pdf')

    return render_template('accounting/client_view.html', customers=customers, customer=customer, customer_id=(customer.id if customer else None), ledger=ledger, deposits=deposits, auction_ledger=auction_ledger, service_rows=service_rows, balances=balances, pl=pl)
